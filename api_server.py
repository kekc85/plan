#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
REST API Backend для электронного журнала смены Диспетчера центровки.
Связывает веб-приложение (React) и парсер AviaBit (Nordwind / Икар).
"""

import os
import sys
import io
import json
import urllib.parse
from datetime import datetime, timezone, timedelta
from typing import List, Optional
from fastapi import FastAPI, HTTPException, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import uvicorn

# Импортируем существующую логику парсера
from parser import (
    AviaBitClient,
    BASE_URL_NORDWIND,
    BASE_URL_IKAR,
    MSK_TZ,
    ALLOWED_DEPARTURES,
    IATA_CITIES,
    parse_date_arg,
    parse_time_arg,
    process_flights
)

app = FastAPI(
    title="AviaBit Shift Log API",
    description="API для синхронизации журнала смены диспетчера с порталом AviaBit",
    version="1.0.0"
)

# Разрешаем CORS для локальной разработки и работы в браузере
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

SHIFT_DATA_FILE = ".current_shift.json"


class FetchScheduleRequest(BaseModel):
    date_from: str
    time_from: str = "08:00"
    date_to: str
    time_to: str = "14:00"
    airline: str = "both"  # "both", "nordwind", "ikar"
    filter_name: str = "WBGarantiya"


class ExportExcelRequest(BaseModel):
    flights: List[dict]
    shift_info: Optional[dict] = None


def calc_release_time_py(flight_time: str, offset_mins: int = 40) -> str:
    """Вычисляет время выпуска (-40 минут)"""
    if not flight_time or ":" not in flight_time:
        return ""
    try:
        parts = flight_time.split(":")
        h = int(parts[0])
        m = int(parts[1])
        total = h * 60 + m - offset_mins
        if total < 0:
            total += 24 * 60
        rel_h = (total // 60) % 24
        rel_m = total % 60
        return f"{rel_h:02d}:{rel_m:02d}"
    except Exception:
        return ""


@app.get("/api/health")
def health_check():
    return {
        "status": "ok",
        "service": "AviaBit Shift Log Backend",
        "time_utc": datetime.now(timezone.utc).isoformat(),
        "time_msk": datetime.now(MSK_TZ).strftime("%H:%M:%S")
    }


@app.post("/api/fetch_schedule")
def fetch_schedule(req: FetchScheduleRequest):
    """
    Парсит расписание с серверов AviaBit за указанный интервал времени (МСК)
    и возвращает форматированный список рейсов для журнала смены.
    """
    try:
        start_date = parse_date_arg(req.date_from)
        end_date = parse_date_arg(req.date_to)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Неверный формат даты: {e}")

    s_h, s_m = parse_time_arg(req.time_from, 8, 0)
    e_h, e_m = parse_time_arg(req.time_to, 14, 0)

    start_dt_msk = datetime(start_date.year, start_date.month, start_date.day, s_h, s_m, 0, tzinfo=MSK_TZ)
    end_dt_msk = datetime(end_date.year, end_date.month, end_date.day, e_h, e_m, 59, tzinfo=MSK_TZ)

    client_nws = AviaBitClient(base_url=BASE_URL_NORDWIND, session_filename=".session.json", name="Nordwind")
    client_ikar = AviaBitClient(base_url=BASE_URL_IKAR, session_filename=".session_ikar.json", name="Икар")

    ok_nws = False
    ok_ikar = False

    if req.airline in ("both", "nordwind"):
        ok_nws = client_nws.login()
    if req.airline in ("both", "ikar"):
        ok_ikar = client_ikar.login()

    if not ok_nws and not ok_ikar:
        raise HTTPException(
            status_code=401,
            detail="Не удалось авторизоваться на серверах AviaBit (Nordwind / Икар). Проверьте сессию и соединение."
        )

    all_flights = []

    if ok_nws:
        try:
            t_id_nws = client_nws.get_template_id(req.filter_name)
            fl_nws = client_nws.fetch_schedule(start_date, end_date, t_id_nws)
            for fl in fl_nws:
                fl["_client"] = client_nws
                all_flights.append(fl)
        except Exception as err:
            print(f"[-] Ошибка загрузки Nordwind: {err}")

    if ok_ikar:
        try:
            t_id_ikar = client_ikar.get_template_id(req.filter_name)
            fl_ikar = client_ikar.fetch_schedule(start_date, end_date, t_id_ikar)
            for fl in fl_ikar:
                fl["_client"] = client_ikar
                all_flights.append(fl)
        except Exception as err:
            print(f"[-] Ошибка загрузки Икар: {err}")

    if not all_flights:
        return {
            "success": True,
            "count": 0,
            "message": "Рейсы за выбранный период не найдены",
            "flights": []
        }

    # Обработка через process_flights
    processed_rows = process_flights(all_flights, start_dt_msk=start_dt_msk, end_dt_msk=end_dt_msk)

    formatted_flights = []
    for idx, r in enumerate(processed_rows):
        route_str = r.get("route", "")
        city = ""
        airports = ""
        if "\n" in route_str:
            parts = route_str.split("\n")
            city = parts[0].strip()
            airports = parts[1].strip()
        else:
            airports = route_str.strip()

        std_time = r.get("std", "")
        if std_time and ":" in std_time:
            hp, mp = std_time.split(":")
            std_time = f"{int(hp):02d}:{int(mp):02d}"

        rel_time = calc_release_time_py(std_time, 40)

        tail = r.get("tail", "").replace("RA-", "").replace("RA", "").replace("-", "").strip()
        if len(tail) > 5:
            tail = tail[:5]

        # Автоматическое определение точной даты рейса (в формате число.месяц, например 25.08)
        f_date = r.get("flight_date") or ""
        if not f_date:
            f_date = start_dt_msk.strftime('%d.%m')
            if std_time and ":" in std_time:
                try:
                    sh_hour = int(std_time.split(":")[0])
                    if sh_hour < start_dt_msk.hour:
                        f_date = end_dt_msk.strftime('%d.%m')
                except Exception:
                    pass

        formatted_flights.append({
            "id": f"fl_aviabit_{idx + 1}_{int(datetime.now().timestamp())}",
            "flight": r.get("flight_no", ""),
            "flight_date": f_date,
            "route_city": city,
            "route_airports": airports,
            "time": std_time,
            "release_time": rel_time,
            "ac_num": tail,
            "ac_config": r.get("layout", ""),
            "pax": r.get("pax_notes", ""),
            "crew": r.get("crew", "2/4/0/0"),
            "fuel_block": "",
            "fuel_trip": "",
            "fuel_taxi": "",
            "dow": "",
            "doi": "",
            "galley": "D",
            "mtow": "",
            "lir_sent": False,
            "cargo": "",
            "mail": "",
            "baggage": "",
            "szv_sent": False,
            "ldm_sent": False,
            "status": "pending",
            "notes": ""
        })

    shift_interval = f"{start_dt_msk.strftime('%d.%m.%Y')} — {end_dt_msk.strftime('%d.%m.%Y')}"

    return {
        "success": True,
        "count": len(formatted_flights),
        "shift_interval": shift_interval,
        "flights": formatted_flights
    }


@app.post("/api/export_excel")
def export_excel_endpoint(req: ExportExcelRequest):
    """
    Генерирует Excel файл суточного плана со 100% точным форматированием и стилями openpyxl из parser.py:
    - 15 колонок
    - Calibri шрифты
    - Тонкие и средние рамки
    - Вертикальный текст СЗВ / ЛДМ (90°)
    - Точные пропорциональные ширины и высоты строк (36pt)
    - Альбомный лист А4, Fit to 1 page wide
    """
    if not req.flights:
        raise HTTPException(status_code=400, detail="Список рейсов пуст")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Суточный план"

    # Заголовок
    date_str = ""
    if req.shift_info:
        date_str = req.shift_info.get("date_interval") or req.shift_info.get("date") or ""
    if not date_str:
        date_str = datetime.now(MSK_TZ).strftime("%d.%m.%Y")

    title_text = f"Суточный план Диспетчера группы центровки  {date_str}"

    # Строка 1: Заголовок
    ws.merge_cells("A1:O1")
    title_cell = ws["A1"]
    title_cell.value = title_text
    title_cell.font = Font(name="Calibri", size=12, bold=True, color="000000")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Строка 2: Шапка таблицы (15 колонок)
    headers = [
        "№ рейса",      # A
        "Маршрут",      # B
        "Время",        # C
        "Номер\nВС",    # D
        "Компано\nвка", # E
        "PAX,\nNOTES",  # F
        "Экипаж\nЛ/Б/И/П", # G
        "Топливо",      # H
        "MTOW",         # I
        "LIR",          # J
        "Груз",         # K
        "Почта",        # L
        "Багаж",        # M
        "СЗВ",          # N
        "ЛДМ"           # O
    ]

    header_font = Font(name="Calibri", size=11, bold=True, color="000000")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    header_border = Border(
        left=Side(style="medium", color="000000"),
        right=Side(style="medium", color="000000"),
        top=Side(style="medium", color="000000"),
        bottom=Side(style="medium", color="000000")
    )

    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    ws.row_dimensions[2].height = 36
    for col_idx, h_text in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=h_text)
        cell.font = header_font
        if col_idx in (14, 15):
            cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90)
        else:
            cell.alignment = header_align
        cell.border = header_border

    # Стили данных
    data_font = Font(name="Calibri", size=11)
    bold_data_font = Font(name="Calibri", size=11, bold=True)
    route_data_font = Font(name="Calibri", size=9)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, f in enumerate(req.flights, start=1):
        row_num = idx + 2
        ws.row_dimensions[row_num].height = 36

        # Маршрут
        route_combined = f.get("route_airports") or ""
        if f.get("route_city"):
            route_combined = f"{f['route_city']}\n{route_combined}"

        # Числовые значения
        raw_tail = str(f.get("ac_num") or "")
        tail_val = int(raw_tail) if raw_tail.isdigit() else raw_tail

        raw_config = str(f.get("ac_config") or "")
        config_val = int(raw_config) if raw_config.isdigit() else raw_config

        raw_pax = str(f.get("pax") or "")
        pax_val = int(raw_pax) if raw_pax.isdigit() else raw_pax

        raw_mtow = str(f.get("mtow") or "")
        mtow_val = int(raw_mtow) if raw_mtow.isdigit() else raw_mtow

        # Топливо
        fuel_parts = []
        if f.get("fuel_block"): fuel_parts.append(f"B:{f['fuel_block']}")
        if f.get("fuel_trip"): fuel_parts.append(f"T:{f['fuel_trip']}")
        if f.get("fuel_taxi"): fuel_parts.append(f"Tx:{f['fuel_taxi']}")
        fuel_combined = " ".join(fuel_parts) if fuel_parts else (f.get("fuel") or "")

        lir_val = "ДА" if f.get("lir_sent") else ""
        szv_val = "ДА" if f.get("szv_sent") else ""
        ldm_val = "ДА" if f.get("ldm_sent") else ""

        row_values = [
            f.get("flight") or "",
            route_combined,
            f.get("time") or "",
            tail_val,
            config_val,
            pax_val,
            f.get("crew") or "",
            fuel_combined,
            mtow_val,
            lir_val,
            f.get("cargo") or "",
            f.get("mail") or "",
            f.get("baggage") or "",
            szv_val,
            ldm_val
        ]

        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            if col_idx == 2:
                cell.font = route_data_font
            elif col_idx == 6:
                cell.font = bold_data_font
            else:
                cell.font = data_font

            cell.alignment = align_center
            cell.border = thin_border

    # Точные ширины колонок
    col_widths = {
        "A": 11.0,  # № рейса
        "B": 14.5,  # Маршрут
        "C": 8.5,   # Время
        "D": 9.5,   # Номер ВС
        "E": 11.0,  # Компановка
        "F": 10.0,  # PAX, NOTES
        "G": 10.0,  # Экипаж
        "H": 14.0,  # Топливо
        "I": 9.0,   # MTOW
        "J": 5.5,   # LIR
        "K": 9.5,   # Груз
        "L": 9.5,   # Почта
        "M": 9.5,   # Багаж
        "N": 4.5,   # СЗВ
        "O": 4.5    # ЛДМ
    }

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Фиксация шапки
    ws.freeze_panes = "A3"

    # Параметры печати: Альбомная ориентация A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.2
    ws.page_margins.bottom = 0.2
    ws.page_margins.header = 0.0
    ws.page_margins.footer = 0.0
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False

    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)

    clean_date = "".join(c for c in date_str if c.isalnum() or c in "._-")
    filename = f"Суточный_план_Диспетчера_{clean_date}.xlsx"
    encoded_filename = urllib.parse.quote(filename)

    return StreamingResponse(
        output_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )


@app.get("/api/get_shift")
def get_shift():
    """Загружает сохраненное состояние смены с сервера"""
    if os.path.exists(SHIFT_DATA_FILE):
        try:
            with open(SHIFT_DATA_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"flights": [], "shiftInfo": None}


@app.post("/api/save_shift")
def save_shift(data: dict = Body(...)):
    """Сохраняет текущее состояние смены на сервере"""
    try:
        with open(SHIFT_DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    print("=" * 65)
    print("   ЗАПУСК REST API СЕРВЕРА AVIABIT SHIFT LOG НА ПОРТУ 8000")
    print("=" * 65)
    uvicorn.run("api_server:app", host="127.0.0.1", port=8000, reload=False)
