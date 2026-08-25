#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Парсер расписания авиакомпании с портала AviaBit (https://aviabit.nordwindairlines.ru/plan-flight)
Выгрузка данных в структурированную таблицу Excel (.xlsx) с фильтрацией по датам и аэропортам вылета.
"""

import sys
import os
import argparse
import json
import xml.etree.ElementTree as ET
from datetime import datetime, timezone, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Московское время (UTC+3)
MSK_TZ = timezone(timedelta(hours=3))

# Разрешенные аэропорты вылета (IATA)
ALLOWED_DEPARTURES = {
    "KQT", "VRA", "GOI", "GOX", "DYU", "ISB", "CCC", "CXR",
    "HOG", "REN", "OSS", "PMW", "ROV", "XIY", "AER", "SUI",
    "UUD", "UTP", "LBD", "HTA", "SSH", "SVO", "TAS", "NMA",
    "TJU", "SKD"
}

# Соответствие кодов аэропортов названиям городов (IATA -> Город)
IATA_CITIES = {
    # Базовые аэропорты
    "KQT": "Бохтар",
    "VRA": "Варадеро",
    "GOI": "Гоа",
    "GOX": "Гоа",
    "DYU": "Дущанбе",
    "ISB": "Исламобад",
    "CCC": "Кайококо",
    "CXR": "Камрань",
    "HOG": "Ольгин",
    "REN": "Оренбург",
    "OSS": "Ош",
    "PMW": "Парламар",
    "PMV": "Парламар",
    "ROV": "Ростов",
    "XIY": "Сиань Сяньян",
    "AER": "Сочи",
    "SUI": "Сухум",
    "UUD": "Улан-Удэ",
    "UTP": "Утапао",
    "LBD": "Худжант",
    "HTA": "Чита",
    "SSH": "Шарм Эль Шейх",
    "SVO": "Москва",
    "TAS": "Ташкент",
    "NMA": "Наманган",
    "TJU": "Куляб",
    "SKD": "Самарканд",

    # Города прилёта
    "KZN": "Казань",
    "BAX": "Барнаул",
    "SCW": "Сыктывкар",
    "LED": "Питер",
    "UFA": "Уфа",
    "KGD": "Калининград",
    "NBC": "Нижнекамск",
    "MQF": "Магнитогорск",
    "CEK": "Челябинск",
    "NOZ": "Новокузнецк",
    "GOJ": "Н.Новгород",
    "KUF": "Самара",
    "OMS": "Омск",
    "OVB": "Новосибирск",
    "SVX": "Екатеринбург",
    "IKT": "Иркутск",
    "KJA": "Красноярск",
    "VVO": "Владивосток",
    "KHV": "Хабаровск",
    "PEE": "Пермь",
    "TOF": "Томск",
    "TJM": "Тюмень",
    "MRV": "Мин.Воды",
    "MCX": "Махачкала",
    "GRV": "Грозный",
    "VOG": "Волгоград",
    "ASF": "Астрахань",
    "IJK": "Ижевск",
    "CSY": "Чебоксары",
    "KRR": "Краснодар",
    "AAQ": "Анапа",
    "AYT": "Анталья",
    "IST": "Стамбул",
    "DXB": "Дубай",
    "DWC": "Дубай",
    "HRG": "Хургада",
    "BHK": "Бухара",
    "FEG": "Фергана",
    "UGU": "Ургенч",
    "FRU": "Бишкек",
    "EVN": "Ереван",
    "GYD": "Баку",
    "TBS": "Тбилиси"
}
DEPARTURE_CITIES = IATA_CITIES

from urllib.parse import urlparse

def _load_env_file():
    """Загрузка переменных окружения из файла .env."""
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if os.path.exists(env_path):
        try:
            with open(env_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#") or "=" not in line:
                        continue
                    k, v = line.split("=", 1)
                    k = k.strip()
                    v = v.strip().strip("'\"")
                    if k and k not in os.environ:
                        os.environ[k] = v
        except Exception:
            pass

_load_env_file()

DEFAULT_USERNAME = os.getenv("AVIABIT_USERNAME", "")
DEFAULT_PASSWORD = os.getenv("AVIABIT_PASSWORD", "")
BASE_URL_NORDWIND = os.getenv("AVIABIT_BASE_URL_NORDWIND", "https://aviabit.nordwindairlines.ru")
BASE_URL_IKAR = os.getenv("AVIABIT_BASE_URL_IKAR", "https://aviabit.ikar.aero")


class AviaBitClient:
    """Клиент для взаимодействия с API AviaBit с сохранением сессии и поддержкой 2FA."""

    def __init__(self, username=DEFAULT_USERNAME, password=DEFAULT_PASSWORD, base_url=BASE_URL_NORDWIND, session_filename=".session.json", name="Nordwind"):
        self.username = username
        self.password = password
        self.base_url = base_url.rstrip("/")
        self.name = name
        self.domain = urlparse(self.base_url).netloc
        self.session_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), session_filename)
        self.session = requests.Session()
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            "Content-Type": "application/json",
            "Origin": self.base_url,
            "Referer": f"{self.base_url}/plan-flight",
            "Accept": "application/json, text/plain, */*"
        }

    def _load_cached_session(self) -> bool:
        """Загрузка сохраненных cookie сессии."""
        if os.path.exists(self.session_file):
            try:
                with open(self.session_file, "r", encoding="utf-8") as f:
                    cookies = json.load(f)
                    for k, v in cookies.items():
                        self.session.cookies.set(k, v, domain=self.domain)
                # Проверка валидности сессии
                test_url = f"{self.base_url}/api/filter-template?code=1001"
                resp = self.session.get(test_url, headers=self.headers, timeout=10)
                if resp.status_code == 200 and isinstance(resp.json(), list):
                    print(f"[+] [{self.name}] Использована действующая сохраненная сессия.")
                    return True
            except Exception:
                pass
        return False

    def _save_session(self):
        """Сохранение cookie сессии на диск."""
        try:
            cookies_dict = self.session.cookies.get_dict()
            with open(self.session_file, "w", encoding="utf-8") as f:
                json.dump(cookies_dict, f)
        except Exception:
            pass

    def login(self, prompt_code_callback=None) -> bool:
        """
        Авторизация в системе с поддержкой 2FA-кода подтверждения из почты.
        """
        if self._load_cached_session():
            return True

        auth_payload = {
            "rememberMe": True,
            "version": {
                "date": "2026-08-06T08:00:00.000Z",
                "company": 'ООО "АвиаБит"',
                "number": "9.8.1"
            },
            "eng": False,
            "username": self.username,
            "password": self.password
        }
        
        url = f"{self.base_url}/api/auth"
        try:
            resp = self.session.post(url, json=auth_payload, headers=self.headers, timeout=20)
            if resp.status_code == 200:
                data = resp.json()
                user = data.get("user", {})
                print(f"[+] [{self.name}] Успешная авторизация: {user.get('name', self.username)} (ID: {user.get('userId')})")
                self._save_session()
                return True
            elif resp.status_code == 401:
                # Требуется 2FA подтверждение по почте
                err_data = {}
                try:
                    err_data = resp.json()
                except Exception:
                    pass

                if err_data.get("code") == "CHECK_AUTH_FAILED":
                    sending_info = err_data.get("info", {}).get("sendingInfo", "почту")
                    print(f"[*] [{self.name}] Требуется код подтверждения (отправлен на {sending_info})")
                    
                    code = None
                    if prompt_code_callback:
                        code = prompt_code_callback(self.name, sending_info)
                    else:
                        code = input(f"[{self.name}] Введите код подтверждения из письма ({sending_info}): ").strip()

                    if not code:
                        print(f"[-] [{self.name}] Код подтверждения не введен.")
                        return False

                    auth_payload["confirmationCode"] = str(code).strip()
                    auth_payload["confirmationType"] = 2

                    resp2 = self.session.post(url, json=auth_payload, headers=self.headers, timeout=20)
                    if resp2.status_code == 200:
                        data2 = resp2.json()
                        user2 = data2.get("user", {})
                        print(f"[+] [{self.name}] Успешная авторизация 2FA: {user2.get('name', self.username)}")
                        self._save_session()
                        return True
                    else:
                        print(f"[-] [{self.name}] Ошибка проверки 2FA кода (код {resp2.status_code}): {resp2.text}")
                        return False
                else:
                    print(f"[-] [{self.name}] Ошибка авторизации 401: {resp.text}")
                    return False
            elif resp.status_code == 429:
                err_data = {}
                try:
                    err_data = resp.json()
                except Exception:
                    pass
                retry = err_data.get("info", {}).get("retryAfter", "несколько")
                print(f"[-] [{self.name}] Лимит попыток входа (429). Повторите через {retry} сек.")
                return False
            else:
                print(f"[-] [{self.name}] Ошибка авторизации (код {resp.status_code}): {resp.text}")
                return False
        except Exception as e:
            print(f"[-] [{self.name}] Исключение при авторизации: {e}")
            return False

    def get_template_id(self, template_name="WBGarantiya") -> int:
        """Получение ID фильтра по названию."""
        url = f"{self.base_url}/api/filter-template?code=1001"
        try:
            resp = self.session.get(url, headers=self.headers, timeout=15)
            if resp.status_code == 200:
                templates = resp.json()
                for item in templates:
                    if item.get("text", "").strip() == template_name:
                        return int(item.get("value"))
                for item in templates:
                    if template_name.lower() in item.get("text", "").lower():
                        return int(item.get("value"))
        except Exception as e:
            print(f"[!] [{self.name}] Не удалось загрузить шаблоны фильтров: {e}")
        
        return 1055

    def fetch_schedule(self, start_dt: datetime, end_dt: datetime, template_id: int = 1055) -> list:
        """
        Запрос расписания полетов за указанный диапазон дат.
        """
        start_bound = datetime(start_dt.year, start_dt.month, start_dt.day, 0, 0, 0)
        end_bound = datetime(end_dt.year, end_dt.month, end_dt.day, 23, 59, 59, 999000)

        ts_start = int(start_bound.timestamp() * 1000)
        ts_end = int(end_bound.timestamp() * 1000)

        url = (
            f"{self.base_url}/api/plan-flight?"
            f"dateBegin={ts_start}&dateEnd={ts_end}&eng=false&apCode=3&apId=0&template={template_id}&showCancel=false"
        )

        print(f"[*] [{self.name}] Запрос расписания с {start_bound.strftime('%d.%m.%Y')} по {end_bound.strftime('%d.%m.%Y')}...")
        try:
            resp = self.session.get(url, headers=self.headers, timeout=30)
            if resp.status_code == 200:
                flights = resp.json()
                print(f"[+] [{self.name}] Получено записей: {len(flights)}")
                return flights
            else:
                print(f"[-] [{self.name}] Ошибка получения расписания (код {resp.status_code}): {resp.text}")
                return []
        except Exception as e:
            print(f"[-] [{self.name}] Ошибка сетевого запроса: {e}")
            return []

    def fetch_flight_preliminary(self, pf_record_id: int) -> dict:
        """
        Запрос оперативной информации по рейсу (пассажиры, топливо, города).
        """
        if not pf_record_id:
            return {}
        url = f"{self.base_url}/api/preliminary-crew-load?planFlightId={pf_record_id}&eng=false"
        try:
            resp = self.session.get(url, headers=self.headers, timeout=5)
            if resp.status_code == 200:
                return resp.json()
        except Exception:
            pass
        return {}


def parse_crew(crew_xml_str: str) -> tuple:
    """
    Парсинг структуры экипажа из XML строки.
    Возвращает кортеж (летный, салон, итс, пассажиры).
    """
    cockpit = 0
    cabin = 0
    its = 0
    pax = 0

    if not crew_xml_str:
        return cockpit, cabin, its, pax

    try:
        root = ET.fromstring(crew_xml_str)
        for emp in root.findall("employee"):
            ctype = emp.get("crewType")
            if ctype == "0":
                cockpit += 1
            elif ctype == "1":
                cabin += 1
            elif ctype == "2":
                its += 1
            elif ctype == "4":
                pax += 1
            else:
                arm = emp.get("armChair", "")
                if arm in ("КС", "2П", "ПИ", "КСи", "КВС/Н"):
                    cockpit += 1
                elif arm in ("СБ", "БП", "БИ"):
                    cabin += 1
                elif arm in ("ИС", "ИТС"):
                    its += 1
                elif arm == "П":
                    pax += 1
    except Exception:
        pass

    return cockpit, cabin, its, pax


def parse_pax_count(pax_str: str) -> str:
    """
    Парсинг количества пассажиров (PAX, NOTES).
    Берется сумма первых двух чисел (взрослые + ребенок большой РБ),
    третья цифра (ребенок маленький РМ / инфант без места) не учитывается.
    """
    if not pax_str:
        return ""
    parts = str(pax_str).strip().split("/")
    if len(parts) >= 2:
        try:
            adults = int(parts[0].strip())
            children = int(parts[1].strip())
            return str(adults + children)
        except ValueError:
            pass
    elif len(parts) == 1:
        try:
            return str(int(parts[0].strip()))
        except ValueError:
            return str(pax_str).strip()
    return str(pax_str).strip()


def process_flights(
    candidate_flights: list,
    preliminaries: dict = None,
    start_dt_msk: datetime = None,
    end_dt_msk: datetime = None
) -> list:
    """
    Обработка и форматирование данных рейсов под итоговую таблицу.
    Фильтрует по диапазону дат и времени (МСК), SVO (pax==0) и сортирует по времени вылета.
    """
    if preliminaries is None:
        preliminaries = {}

    # Дедупликация и фильтрация рейсов (исключаем резервы ~РЕЗ с красной буквой R)
    seen_keys = set()
    unique_candidates = []
    for fl in candidate_flights:
        flight_no = (fl.get("flight") or "").strip()
        dep = (fl.get("airPortTOCode") or "").strip().upper()
        arr = (fl.get("airPortLACode") or "").strip().upper()

        # Исключаем резервы (~РЕЗ, ~РЕЗ-04 и т.д.), спецзаписи с буквой R и пустые номера
        if not flight_no or flight_no.startswith("~") or "~" in flight_no:
            continue
        if "РЕЗ" in flight_no.upper() or "REZ" in flight_no.upper():
            continue
        if fl.get("isSpecialFlight") is True:
            continue

        # Фильтр по разрешенным аэропортам вылета
        if dep not in ALLOWED_DEPARTURES:
            continue

        fl_key = (
            flight_no.replace("-", "").replace(" ", "").upper(),
            fl.get("dateTakeoff"),
            dep,
            arr
        )
        if fl_key not in seen_keys:
            seen_keys.add(fl_key)
            unique_candidates.append(fl)

    # Параллельная дозагрузка preliminary (пассажиры/экипаж/города), если не была загружена
    missing_pf_ids = [
        (fl.get("pfRecordId"), fl.get("_client"))
        for fl in unique_candidates
        if fl.get("pfRecordId") and fl.get("pfRecordId") not in preliminaries
    ]

    if missing_pf_ids:
        print(f"[*] Загрузка оперативной информации (пассажиры/продажи) для {len(missing_pf_ids)} рейсов...")
        with ThreadPoolExecutor(max_workers=30) as executor:
            future_to_id = {
                executor.submit(client.fetch_flight_preliminary, pf_id): pf_id
                for pf_id, client in missing_pf_ids
                if client is not None
            }
            for future in as_completed(future_to_id):
                pf_id = future_to_id[future]
                try:
                    preliminaries[pf_id] = future.result()
                except Exception:
                    preliminaries[pf_id] = {}

    processed = []

    for fl in unique_candidates:
        flight_no = (fl.get("flight") or "").strip()
        dep = (fl.get("airPortTOCode") or "").strip().upper()
        arr = (fl.get("airPortLACode") or "").strip().upper()
        pf_id = fl.get("pfRecordId")

        # Номер рейса без дефисов (например N4-1442 -> N41442, EO-487 -> EO487)
        flight_clean = flight_no.replace("-", "").replace(" ", "")

        # Парсим актуальное время вылета (ATD/ETD с учетом задержек)
        # Приоритет: 1. Фактическое (dateTakeoffReal) -> 2. Расчетное с задержкой (dateTakeoffCalculation) -> 3. Плановое (dateTakeoff)
        takeoff_raw = fl.get("dateTakeoffReal") or fl.get("dateTakeoffCalculation") or fl.get("dateTakeoff")
        time_str = ""
        sort_timestamp = 0
        dt_msk = None

        if takeoff_raw:
            try:
                dt_utc = datetime.fromisoformat(takeoff_raw.replace("Z", "+00:00"))
                # Перевод в Московское время (МСК = UTC+3)
                dt_msk = dt_utc.astimezone(MSK_TZ)
                # Формат H:MM (например 7:20, 1:35, 11:15)
                time_str = f"{dt_msk.hour}:{dt_msk.strftime('%M')}"
                sort_timestamp = int(dt_msk.timestamp())
            except Exception:
                pass

        # Фильтрация по точному диапазону времени в Московском часовом поясе
        if start_dt_msk and end_dt_msk and dt_msk:
            if dt_msk < start_dt_msk or dt_msk > end_dt_msk:
                continue

        # Номер ВС (очищенный от префикса RA)
        raw_tail = (fl.get("pln") or "").strip()
        tail_clean = raw_tail.replace("RA-", "").replace("RA", "").replace("-", "").strip()

        # Компановка
        layout = (fl.get("prePlaneComponovkaInfo") or "").strip()

        # Экипаж (4 разные цифры через слеш: Летный/Салон/ИТС/Пассажиры)
        cockpit, cabin, its, pax = parse_crew(fl.get("crew"))
        crew_str = f"{cockpit}/{cabin}/{its}/{pax}"

        # Оперативная информация (пассажиры/продажи)
        pax_raw = ""

        oper_data = preliminaries.get(pf_id, {})
        preliminary_list = oper_data.get("preliminary", [])
        if preliminary_list and isinstance(preliminary_list, list):
            leg0 = preliminary_list[0]
            if isinstance(leg0, dict):
                pax_raw = (leg0.get("prePassengerInfo") or "").strip()
                # Если компоновка в основном списке была пустой, берем из preliminary
                if not layout:
                    layout = (leg0.get("prePlaneComponovkaInfo") or "").strip()

        pax_notes = parse_pax_count(pax_raw)

        # Условие для SVO: рейсы с вылетом из SVO включаются ТОЛЬКО если пассажиров 0 (пустые/перегоночные)
        if dep == "SVO":
            pax_int = int(pax_notes) if str(pax_notes).isdigit() else 0
            if pax_notes and pax_int > 0:
                continue

        # Маршрут: город ПРИЛЁТА сверху, DEP-ARR снизу
        city_name = IATA_CITIES.get(arr) or ""
        if not city_name and preliminary_list and isinstance(preliminary_list, list):
            # Если аэропорта нет в словаре, парсим динамически из 2-го сегмента ответа сервера
            if len(preliminary_list) > 1 and isinstance(preliminary_list[1], dict):
                raw_name = (preliminary_list[1].get("airportName") or "").strip()
                if "(" in raw_name:
                    raw_name = raw_name.split("(")[0].strip()
                if raw_name:
                    city_name = raw_name

        if city_name:
            route_str = f"{city_name}\n{dep}-{arr}"
        else:
            route_str = f"{dep}-{arr}"

        processed.append({
            "sort_ts": sort_timestamp,
            "flight_no": flight_clean,
            "flight_date": dt_msk.strftime("%d.%m") if dt_msk else "",
            "route": route_str,
            "std": time_str,
            "tail": tail_clean,
            "layout": layout,
            "pax_notes": pax_notes,
            "crew": crew_str,
            "fuel": "",  # Ячейка топлива остается пустой для ручного ввода
            "dep": dep,
            "arr": arr
        })

    # Хронологическая сортировка по времени вылета STD
    processed.sort(key=lambda x: (x["sort_ts"], x["dep"], x["flight_no"]))
    return processed


def export_to_excel(
    rows: list,
    output_filename: str,
    start_dt: datetime,
    end_dt: datetime
):
    """
    Экспорт данных в таблицу Excel по точному формату "Суточный план Диспетчера группы центровки".
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Суточный план"

    # Форматирование диапазона дат и времени для заголовка
    if start_dt.hour == 0 and start_dt.minute == 0 and end_dt.hour == 23 and end_dt.minute >= 59:
        start_fmt = start_dt.strftime("%d.%m.")
        end_fmt = end_dt.strftime("%d.%m.%y")
        date_header_str = f"{start_fmt}-{end_fmt}"
    else:
        start_fmt = start_dt.strftime("%d.%m.%y %H:%M")
        end_fmt = end_dt.strftime("%d.%m.%y %H:%M")
        date_header_str = f"{start_fmt} - {end_fmt}"

    title_text = f"Суточный план Диспетчера группы центровки  {date_header_str}"

    # Строка 1: Главный заголовок
    ws.merge_cells("A1:O1")
    title_cell = ws["A1"]
    title_cell.value = title_text
    title_cell.font = Font(name="Calibri", size=12, bold=True, color="000000")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Шапка таблицы (15 колонок)
    headers = [
        "№ рейса",      # A
        "Маршрут",      # B
        "Время",        # C
        "Номер\nВС",    # D
        "Компано\nвка", # E
        "PAX,\nNOTES",  # F
        "Экипаж\nЛ/Б/И/П", # G
        "Топливо",      # H
        "MTOW",         # I
        "LIR",          # J
        "Груз",         # K
        "Почта",        # L
        "Багаж",        # M
        "СЗВ",          # N
        "ЛДМ"           # O
    ]

    header_font = Font(name="Calibri", size=11, bold=True, color="000000")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    header_border = Border(
        left=Side(style="medium", color="000000"),
        right=Side(style="medium", color="000000"),
        top=Side(style="medium", color="000000"),
        bottom=Side(style="medium", color="000000")
    )

    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    ws.row_dimensions[2].height = 36
    for col_idx, h_text in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=h_text)
        cell.font = header_font
        if col_idx in (14, 15):
            # Вертикальная ориентация текста для колонок СЗВ и ЛДМ
            cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90)
        else:
            cell.alignment = header_align
        cell.border = header_border

    # Стили строк данных (Calibri 11 для таблицы, Calibri 9 для маршрута с городами)
    data_font = Font(name="Calibri", size=11)
    bold_data_font = Font(name="Calibri", size=11, bold=True)
    route_data_font = Font(name="Calibri", size=9)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, r_data in enumerate(rows, start=1):
        row_num = idx + 2
        ws.row_dimensions[row_num].height = 36

        # Преобразование числовых полей в int для исключения зеленых уголков (число как текст)
        tail_val = int(r_data["tail"]) if str(r_data.get("tail", "")).isdigit() else r_data.get("tail", "")
        layout_val = int(r_data["layout"]) if str(r_data.get("layout", "")).isdigit() else r_data.get("layout", "")
        pax_val = int(r_data["pax_notes"]) if str(r_data.get("pax_notes", "")).isdigit() else r_data.get("pax_notes", "")

        row_values = [
            r_data["flight_no"],   # A: № рейса
            r_data["route"],       # B: Маршрут
            r_data["std"],         # C: Время
            tail_val,              # D: Номер ВС (число)
            layout_val,            # E: Компановка (число)
            pax_val,               # F: PAX, NOTES (число)
            r_data["crew"],        # G: Экипаж
            r_data["fuel"],        # H: Топливо
            "",                    # I: MTOW
            "",                    # J: LIR
            "",                    # K: Груз
            "",                    # L: Почта
            "",                    # M: Багаж
            "",                    # N: СЗВ
            ""                     # O: ЛДМ
        ]

        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            # Шрифт ячейки: 9 для Маршрута (B), 11 Bold для PAX (F), 11 обычный для остальных (включая рейс)
            if col_idx == 2:
                cell.font = route_data_font
            elif col_idx == 6:
                cell.font = bold_data_font
            else:
                cell.font = data_font

            cell.alignment = align_center
            cell.border = thin_border

    # Настройка пропорциональных ширин колонок для полного заполнения листа А4 Альбом
    col_widths = {
        "A": 11.0,  # № рейса
        "B": 14.5,  # Маршрут
        "C": 8.5,   # Время
        "D": 9.5,   # Номер ВС
        "E": 11.0,  # Компановка
        "F": 10.0,  # PAX, NOTES
        "G": 10.0,  # Экипаж
        "H": 14.0,  # Топливо
        "I": 9.0,   # MTOW
        "J": 5.5,   # LIR
        "K": 9.5,   # Груз
        "L": 9.5,   # Почта
        "M": 9.5,   # Багаж
        "N": 4.5,   # СЗВ (вертикально)
        "O": 4.5    # ЛДМ (вертикально)
    }

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Фиксация шапки таблицы
    ws.freeze_panes = "A3"

    # Параметры печати: Альбомная ориентация A4 с минимальными полями как на образце
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.2
    ws.page_margins.bottom = 0.2
    ws.page_margins.header = 0.0
    ws.page_margins.footer = 0.0
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False

    # Сохранение файла с обработкой блокировки (если файл открыт в Excel)
    actual_path = output_filename
    try:
        wb.save(output_filename)
    except PermissionError:
        base, ext = os.path.splitext(output_filename)
        for i in range(1, 100):
            alt_path = f"{base}_({i}){ext}"
            try:
                wb.save(alt_path)
                actual_path = alt_path
                print(f"[!] Исходный файл открыт в Excel. Сохранено в: {actual_path}")
                break
            except PermissionError:
                continue

    print(f"[+] Файл успешно сохранен: {os.path.abspath(actual_path)}")
    return actual_path


def parse_time_arg(time_str: str, default_h: int = 0, default_m: int = 0) -> tuple[int, int]:
    """Парсинг строкового времени в формате ЧЧ:ММ."""
    if not time_str:
        return default_h, default_m
    time_str = time_str.strip().replace("-", ":").replace(".", ":")
    try:
        parts = time_str.split(":")
        if len(parts) == 1:
            h = int(parts[0])
            m = 0
        elif len(parts) >= 2:
            h = int(parts[0])
            m = int(parts[1])
        else:
            return default_h, default_m
        if 0 <= h <= 23 and 0 <= m <= 59:
            return h, m
    except Exception:
        pass
    return default_h, default_m


def parse_date_arg(date_str: str) -> datetime:
    """Парсинг строковой даты (поддерживает ДД.ММ.ГГГГ, ДД/ММ/ГГГГ и сплошной ввод ДДММГГГГ)."""
    date_str = date_str.strip()
    digits = "".join(c for c in date_str if c.isdigit())
    if len(digits) == 8:
        try:
            return datetime.strptime(digits, "%d%m%Y")
        except ValueError:
            pass
    elif len(digits) == 6:
        try:
            return datetime.strptime(digits, "%d%m%y")
        except ValueError:
            pass

    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            pass
    try:
        cur_year = datetime.now().year
        return datetime.strptime(f"{date_str}.{cur_year}", "%d.%m.%Y")
    except ValueError:
        pass
    raise ValueError(f"Неверный формат даты: {date_str}. Используйте ДД.ММ.ГГГГ")


def run_parse(
    start_date: datetime,
    end_date: datetime,
    output_path: str = None,
    filter_name: str = "WBGarantiya",
    prompt_code_callback=None,
    start_time_str: str = "00:00",
    end_time_str: str = "23:59"
) -> tuple[bool, str, int, str]:
    """
    Основная программная функция выгрузки расписания с учетом диапазона дат и времени (МСК).
    Выгружает данные с Nordwind и Икар, объединяет и сортирует хронологически.
    Возвращает (успех, сообщение, количество_рейсов, путь_к_файлу).
    """
    s_h, s_m = parse_time_arg(start_time_str, 0, 0)
    e_h, e_m = parse_time_arg(end_time_str, 23, 59)

    start_dt_msk = datetime(start_date.year, start_date.month, start_date.day, s_h, s_m, 0, tzinfo=MSK_TZ)
    end_dt_msk = datetime(end_date.year, end_date.month, end_date.day, e_h, e_m, 59, tzinfo=MSK_TZ)

    start_str = start_dt_msk.strftime("%d.%m.%Y")
    end_str = end_dt_msk.strftime("%d.%m.%Y")

    if not output_path:
        if s_h == 0 and s_m == 0 and e_h == 23 and e_m == 59:
            output_path = f"Суточный_план_Диспетчера_{start_str}_{end_str}.xlsx"
        else:
            s_t = f"{s_h:02d}-{s_m:02d}"
            e_t = f"{e_h:02d}-{e_m:02d}"
            output_path = f"Суточный_план_Диспетчера_{start_str}_{s_t}_{end_str}_{e_t}.xlsx"

    client_nws = AviaBitClient(base_url=BASE_URL_NORDWIND, session_filename=".session.json", name="Nordwind")
    client_ikar = AviaBitClient(base_url=BASE_URL_IKAR, session_filename=".session_ikar.json", name="Икар")

    ok_nws = client_nws.login(prompt_code_callback=prompt_code_callback)
    ok_ikar = client_ikar.login(prompt_code_callback=prompt_code_callback)

    if not ok_nws and not ok_ikar:
        return False, "Не удалось авторизоваться ни на одном из порталов (Nordwind / Икар). Проверьте доступ и коды 2FA.", 0, ""

    all_flights = []

    if ok_nws:
        t_id_nws = client_nws.get_template_id(filter_name)
        fl_nws = client_nws.fetch_schedule(start_date, end_date, t_id_nws)
        for fl in fl_nws:
            fl["_client"] = client_nws
            all_flights.append(fl)

    if ok_ikar:
        t_id_ikar = client_ikar.get_template_id(filter_name)
        fl_ikar = client_ikar.fetch_schedule(start_date, end_date, t_id_ikar)
        for fl in fl_ikar:
            fl["_client"] = client_ikar
            all_flights.append(fl)

    if not all_flights:
        return False, f"Рейсы за период {start_str} - {end_str} не найдены ни на одном сервере.", 0, ""

    rows = process_flights(all_flights, start_dt_msk=start_dt_msk, end_dt_msk=end_dt_msk)
    if not rows:
        return False, "Нет рейсов, соответствующих указанным аэропортам вылета и интервалу времени.", 0, ""

    actual_saved_path = export_to_excel(rows, output_path, start_dt_msk, end_dt_msk)
    return True, f"Успешно выгружено {len(rows)} рейсов (Nordwind + Икар) в файл:\n{os.path.abspath(actual_saved_path)}", len(rows), actual_saved_path


def main():
    parser = argparse.ArgumentParser(description="Парсер расписания полетов AviaBit в Excel.")
    parser.add_argument("--start", "-s", type=str, help="Дата начала периода (ДД.ММ.ГГГГ)")
    parser.add_argument("--end", "-e", type=str, help="Дата окончания периода (ДД.ММ.ГГГГ)")
    parser.add_argument("--out", "-o", type=str, help="Путь к выходному файлу Excel (.xlsx)")
    parser.add_argument("--filter", "-f", type=str, default="WBGarantiya", help="Название фильтра (по умолчанию WBGarantiya)")
    args = parser.parse_args()

    print("=" * 65)
    print("        ПАРСЕР РАСПИСАНИЯ ПОЛЕТОВ AVIABIT -> EXCEL")
    print("=" * 65)

    start_date = None
    end_date = None

    if args.start:
        try:
            start_date = parse_date_arg(args.start)
        except ValueError as err:
            print(f"[-] {err}")
            sys.exit(1)
    else:
        while not start_date:
            def_start = datetime.now().strftime("%d.%m.%Y")
            val = input(f"Введите дату начала [ДД.ММ.ГГГГ] (по умолчанию {def_start}): ").strip()
            if not val:
                val = def_start
            try:
                start_date = parse_date_arg(val)
            except ValueError as err:
                print(f"[-] {err}")

    if args.end:
        try:
            end_date = parse_date_arg(args.end)
        except ValueError as err:
            print(f"[-] {err}")
            sys.exit(1)
    else:
        while not end_date:
            def_end = start_date.strftime("%d.%m.%Y")
            val = input(f"Введите дату окончания [ДД.ММ.ГГГГ] (по умолчанию {def_end}): ").strip()
            if not val:
                val = def_end
            try:
                end_date = parse_date_arg(val)
                if end_date < start_date:
                    print("[-] Дата окончания не может быть раньше даты начала!")
                    end_date = None
            except ValueError as err:
                print(f"[-] {err}")

    success, msg, count, saved_file = run_parse(start_date, end_date, args.out, args.filter)
    if success:
        print("=" * 65)
        print(msg)
        print("=" * 65)
    else:
        print(f"[-] {msg}")
        sys.exit(1)


if __name__ == "__main__":
    main()
